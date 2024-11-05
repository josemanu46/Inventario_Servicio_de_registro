import sys
import os
import time
import pdb
from tkinter import messagebox
import numpy as np
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
import logging
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill
import linecache


def my_custom_logger(self, logger_name, level=logging.DEBUG,format_string='%(levelname)s:%(asctime)s:%(message)s'):
    """
    Method to return a custom logger with the given name and level
    """
    logger = logging.getLogger(logger_name)
    logger.setLevel(level)
    log_format = logging.Formatter(format_string)
    # Creating and adding the console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(log_format)
    logger.addHandler(console_handler)
    # Creating and adding the file handler
    file_handler = logging.FileHandler(logger_name, mode='a')
    file_handler.setFormatter(log_format)
    logger.addHandler(file_handler)
    return logger

def PrintException(self):
    exc_type, exc_obj, tb = sys.exc_info()
    f = tb.tb_frame
    lineno = tb.tb_lineno
    filename = f.f_code.co_filename
    linecache.checkcache(filename)
    line = linecache.getline(filename, lineno, f.f_globals)

    salida = 'EXCEPTION IN LINE {} "{}": {}'.format(lineno, line.strip(), exc_obj)

    self.logger.info(salida)

def generate_report(card_report,sfp_report,frame_report,atp_Inventario):
    print('Start Generate report...')
    df_card_full = pd.read_excel(card_report, sheet_name=0,header=0)
    df_sfp_full = pd.read_excel(sfp_report, sheet_name= 0, header=0)
    df_frame_full = pd.read_excel(frame_report, sheet_name= 0, header=0)

    df_card = pd.read_excel(card_report, sheet_name=0,header=3)
    df_sfp = pd.read_excel(sfp_report, sheet_name= 0, header=7)
    df_frame = pd.read_excel(frame_report, sheet_name= 0, header=3)
    df_inventario = pd.read_excel(atp_Inventario, sheet_name=0)
    #sheet_name='Sheet1'
    
    #Eliminar espacios en blanco al principio y al final de las cadenas
    df_inventario = df_inventario.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    #Convertir todas las cadenas a mayúsculas
    df_inventario = df_inventario.apply(lambda x: x.str.upper() if x.dtype == "object" else x)
    # y deseas seleccionar la segunda columna (posición 1 en base a 0)
    #df_inventario
    segunda_columna = df_inventario.iloc[:, 1]
    primera_columna = df_inventario.iloc[:, 0]
    print(segunda_columna)
    print(primera_columna)
    pdb.set_trace()

    try:
        if pd.notna(segunda_columna).all():
            #df_inventario[segunda_columna]= df_inventario[segunda_columna].str.replace('-', '')
            #segunda_columna.str.replace('-', '')
            #df_inventariounda_columna]= df_inventario[segunda_columna].str.replace(' ', '')
            segunda_columna.str.replace(' ', '')
            print('Con: No. de Serie')      
            print(segunda_columna)
            
            pass
        else:
            raise ValueError("Los valores de la segunda columna son NaN")
        #df_inventario = segunda_columna.apply(lambda x: x.str.upper() if x.dtype == "object" else x)
    except ValueError as e:
        print("Error:", e)
        #pdb.set_trace()
        #AQUI OBTENER LOS VALORES DE LAS COLUMNAS 
        #df_inventario[primera_columna]= df_inventario[primeclsra_columna].str.replace('-', '')
        # Descombinar la columna
        df_descombinado = primera_columna.copy()
        #df_descombinado.iloc[:, 0]
        #df_inventario[primera_columna]= df_inventario[primera_columna].str.replace(' ', '')
        df_descombinado.str.replace(' ', '')
        print('Con: Serie')
        # Convertir todas las cadenas a mayúsculas
        #df_inventario = primera_columna.apply(lambda x: x.str.upper() if x.dtype == "object" else x)

        # Encuentra la posición de la columna en la que deseas descombinar
        #posicion_columna = df.columns.get_loc(nombre_columna_combinada)



    print('limpio inventario...')
    
    # ==================== SFP ====================
    # Eliminar espacios en blanco al principio y al final de las cadenas
    df_sfp = df_sfp.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    # Eliminar caracteres específicos (por ejemplo, guiones)
    #df_sfp['SN(Bar Code)']= df_sfp['SN(Bar Code)'].str.replace('-', '')
    df_sfp['SN(Bar Code)']= df_sfp['SN(Bar Code)'].str.replace(' ', '')

    # Convertir todas las cadenas a mayúsculas
    df_sfp = df_sfp.apply(lambda x: x.str.upper() if x.dtype == "object" else x)

    print('limpio SFP...')

    # ==================== CARD ====================
    # Eliminar espacios en blanco al principio y al final de las cadenas
    df_card = df_card.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    # Eliminar caracteres específicos (por ejemplo, guiones)
    #df_card['SN(Bar Code)']= df_card['SN(Bar Code)'].str.replace('-', '')
    df_card['SN(Bar Code)']= df_card['SN(Bar Code)'].str.replace(' ', '')
    # Convertir todas las cadenas a mayúsculas
    df_card = df_card.apply(lambda x: x.str.upper() if x.dtype == "object" else x)

    print('limpio Card...')
    #pdb.set_trace()
    #==================== FRAME ====================
    # Eliminar espacios en blanco al principio y al final de las cadenas
    df_frame = df_frame.apply(lambda x: x.str.strip() if x.dtype == "object" else x)

    # Eliminar caracteres específicos (por ejemplo, guiones)
    #df_card['SN(Bar Code)']= df_card['SN(Bar Code)'].str.replace('-', '')
    df_frame['SN(Bar Code)']= df_frame['SN(Bar Code)'].str.replace(' ', '')
    # Convertir todas las cadenas a mayúsculas
    df_frame = df_frame.apply(lambda x: x.str.upper() if x.dtype == "object" else x)

    print('limpio frame...')



    # eliminar nan, 
    df_card_SN = df_card.dropna(subset=['SN(Bar Code)'])
    df_sfp_SN = df_sfp.dropna(subset=['SN(Bar Code)'])
    df_frame_SN = df_frame.dropna(subset=['SN(Bar Code)'])

    print('Eliminar valores NAN en card, sfp y frame')
    #df_inventario.to_excel('archivo_resultado.xlsx', index=False)
    print('empezando convinacion...')
    # convinar df_card_SN, df_sfp_SN columna SN(Bar Code)
    df_SN_Final = pd.concat([df_card_SN['SN(Bar Code)'], df_sfp_SN['SN(Bar Code)']], axis=0)
    #pdb.set_trace()
    # eliminar duplicados , si es que hay 
    df_SN_Final = df_SN_Final.drop_duplicates()
    #pdb.set_trace()
    df_SN_Final = df_SN_Final.reset_index(drop=True)
    print('df with SN Bar code')
    print(df_SN_Final)


    try: 
        if pd.notna(segunda_columna).all():
            inv = segunda_columna.reset_index(drop=True)
            coincidentes = inv.isin(df_SN_Final)
            df_coincidentes = coincidentes.to_frame()
            print(df_coincidentes)
            df_inv = inv.to_frame()
            print(df_inv)
            print('Metodo 1')
            #print(columnas_a_colorear)
            #aqui ver que pedo con el nombre de salida
            archivo_excel = 'C:\\Users\\j84319062\\Documents\\GitHub\\ATP_Inventario\\Template\\InventarioATP.xlsx'
            template = str(os.path.dirname(os.path.realpath(__file__))) + r"\Template"
            archivo_excel = str(os.path.dirname(os.path.realpath(__file__))) + r"\Template\InventarioATP.xlsx"
            wb = openpyxl.load_workbook(atp_Inventario)
            ws = wb.active
            # Descombinar todas las celdas en la hoja
            #obtener el formato de relleno de color del DataFrame resultado
            fill_format = {
                True: PatternFill(start_color='0EE1A4', end_color='0EE1A4', fill_type='solid'),  #verde si hay 
                False: PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'), #rojo si no hay
            }

            for row_idx, row in enumerate(df_inv.values,start=2):
                for col_idx, value in enumerate(row):
                    cell = ws.cell(row=row_idx, column=col_idx + 2)
                    fill = fill_format.get(df_coincidentes.iat[row_idx-2, col_idx])  # Obtener el formato de relleno de color basado en df
                    if fill:
                        cell.fill = fill
                    
                    #ificar si el valor comienza con "2021" y aplicar otro formato de relleno ci cierto
                    if str(value).startswith("2102"):
                        cell.fill = PatternFill(start_color='1900ff', end_color='1900ff', fill_type='solid') #azul racks
                    
                    cell.value = value  # Asignar el valor a la celda
                    #ws.cell(row=row_idx, column=col_idx+ 2,value=value)
            #ver la forma de copiar los valores con colores en la columna especifica No. de Serie , sobre el archivo Template
            wb.save(archivo_excel)
            print("Archivo guardado exitosamente :D")
            #messagebox.showinfo("ATP Inventario","Los Archivos se han creado exitosamente") 
            #abrir_ubicacion(template)

            pass
        else:
            raise ValueError("Los valores de la segunda columna son NaN")
        #df_inventario = segunda_columna.apply(lambda x: x.str.upper() if x.dtype == "object" else x)
    except ValueError as e:
        print("Error:", e)
        
        inv = df_descombinado.reset_index(drop=True)
        coincidentes = inv.isin(df_SN_Final)
        df_coincidentes = coincidentes.to_frame()
        print(df_coincidentes)
        
        df_inv = inv.to_frame()
        print(df_inv)   
        print('Metodo 2')     
        

        coincidentesRacks = inv.isin(df_frame_SN)
        df_coincidentesRacks = coincidentesRacks.to_frame()
        print(df_coincidentesRacks)
        #print(columnas_a_colorear)
        #aqui ver que pedo con el nombre de salida
        archivo_excel = str(os.path.dirname(os.path.realpath(__file__))) + r"\Template\InventarioATP.xlsx"
        wb = openpyxl.load_workbook(atp_Inventario)
        ws = wb.active


        # Descombinar todas las celdas en la hoja

        #obtener el formato de relleno de color del DataFrame resultado
        fill_format = {
            True: PatternFill(start_color='0EE1A4', end_color='0EE1A4', fill_type='solid'),  #verde si hay 
            False: PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'), #rojo si no hay
        }

        for row_idx, row in enumerate(df_inv.values,start=2):
            for col_idx, value in enumerate(row):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                fill = fill_format.get(df_coincidentes.iat[row_idx-2, col_idx])  # Obtener el formato de relleno de color basado en df
                if fill:
                    cell.fill = fill
                
                if value in df_frame_SN.values and value.startswith('2102'):
                    cell.fill = PatternFill(start_color='1900ff', end_color='1900ff', fill_type='solid') #azul racks

                cell.value = value  # Asignar el valor a la celda
                #ws.cell(row=row_idx, column=col_idx+ 2,value=value)

       
       
       
        #ver la forma de copiar los valores con colores en la columna especifica No. de Serie , sobre el archivo Template
        wb.save(archivo_excel)
        print("Archivo guardado exitosamente :D")
        #messagebox.showinfo("ATP Inventario","Los Archivos se han creado exitosamente") 
        #abrir_ubicacion(template)   
        
        
    print('call funcion :')
    generate_sfp_outputfiles(df_sfp_full,df_sfp,df_descombinado)
    generate_card_outputfiles(df_card_full,df_card,df_descombinado)
    generate_frame_outputfiles(df_frame_full,df_frame,df_descombinado)
    #pdb.set_trace()`


    #Es para probar el reporte de registros vacios
    #pdb.set_trace()


def generate_sfp_outputfiles(df_sfp_full,sfp,inv):
    print('empezar con archivos sfp, card, frame')
    pdb.set_trace()

    archivo_sfp_excel = str(os.path.dirname(os.path.realpath(__file__))) + r"\Template\SFP.xlsx"
    df_sfp_full.to_excel(archivo_sfp_excel,index=False)

    print('creando archivo excel')
    
    alexistexas= sfp['SN(Bar Code)']
    pdb.set_trace()
    print(alexistexas)
    
    dfinv1 = inv.reset_index(drop=True)
    coincidentes = alexistexas.isin(dfinv1)
    df_coincidentes = coincidentes.to_frame()
    print(df_coincidentes)
    df_alexistexas = alexistexas.to_frame()

    if isinstance(df_alexistexas, pd.DataFrame):
        print('es df')
    else:
        print('no es df')
    if isinstance(df_coincidentes, pd.DataFrame):
        print('es df')
    else:
        print('no es df')

    print(df_alexistexas)
    
    wb = openpyxl.load_workbook(archivo_sfp_excel)
    ws = wb.active
    # Descombinar todas las celdas en la hoja
    #obtener el formato de relleno de color del DataFrame resultado
    fill_format = {
        True: PatternFill(start_color='0EE1A4', end_color='0EE1A4', fill_type='solid'),  #verde si hay 
        False: PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'), #rojo si no hay
    }
    for row_idx, row in enumerate(df_alexistexas.values,start=2): #fila
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=row_idx + 7, column=col_idx + 6)
            fill = fill_format.get(df_coincidentes.iat[row_idx -2, col_idx])  # Obtener el formato de relleno de color basado en df
            if fill :
                cell.fill = fill
            cell.value = value
            # Asignar el valor a la celda
            #ws.cell(row=row_idx, column=col_idx+ 2,value=value)
    #ver la forma de copiar los valores con colores en la columna especifica No. de Serie , sobre el archivo Template
    wb.save(archivo_sfp_excel)
    print("Archivo SFP guardado exitosamente :D")
    #messagebox.showinfo("ATP Inventario","Los Archivos se han creado exitosamente") 
    #abrir_ubicacion(template)    


def generate_card_outputfiles(df_card_full,card,inv):
    print('empezar con archivos card, card, frame')
    pdb.set_trace()
    archivo_sfp_excel = str(os.path.dirname(os.path.realpath(__file__))) + r"\Template\Card.xlsx"
    df_card_full.to_excel(archivo_sfp_excel,index=False)

    print('creando archivo excel')
    
    alexistexas= card['SN(Bar Code)']
    pdb.set_trace()
    print(alexistexas)
    
    dfinv1 = inv.reset_index(drop=True)
    coincidentes = alexistexas.isin(dfinv1)
    df_coincidentes = coincidentes.to_frame()
    print(df_coincidentes)
    df_alexistexas = alexistexas.to_frame()

    if isinstance(df_alexistexas, pd.DataFrame):
        print('es df')
    else:
        print('no es df')
    if isinstance(df_coincidentes, pd.DataFrame):
        print('es df')
    else:
        print('no es df')

    print(df_alexistexas)
    pdb.set_trace()
    wb = openpyxl.load_workbook(archivo_sfp_excel)
    ws = wb.active
    # Descombinar todas las celdas en la hoja
    #obtener el formato de relleno de color del DataFrame resultado
    fill_format = {
        True: PatternFill(start_color='0EE1A4', end_color='0EE1A4', fill_type='solid'),  #verde si hay 
        False: PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'), #rojo si no hay
    }
    for row_idx, row in enumerate(df_alexistexas.values,start=2): #fila
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=row_idx + 3, column=col_idx + 13)
            fill = fill_format.get(df_coincidentes.iat[row_idx -2, col_idx])  # Obtener el formato de relleno de color basado en df
            if fill :
                cell.fill = fill
            cell.value = value
            # Asignar el valor a la celda
            #ws.cell(row=row_idx, column=col_idx+ 2,value=value)
    #ver la forma de copiar los valores con colores en la columna especifica No. de Serie , sobre el archivo Template
    wb.save(archivo_sfp_excel)
    print("Archivo CARD guardado exitosamente :D")
    #messagebox.showinfo("ATP Inventario","Los Archivos se han creado exitosamente") 
    #abrir_ubicacion(template)    


def generate_frame_outputfiles(df_frame_full,frame,inv):
    print('empezar con archivos card, card, frame')
    pdb.set_trace()
    archivo_sfp_excel = str(os.path.dirname(os.path.realpath(__file__))) + r"\Template\Frame.xlsx"
    df_frame_full.to_excel(archivo_sfp_excel,index=False)

    print('creando archivo excel')
    
    alexistexas= frame['SN(Bar Code)']
    pdb.set_trace()
    print(alexistexas)
    
    dfinv1 = inv.reset_index(drop=True)
    coincidentes = alexistexas.isin(dfinv1)
    df_coincidentes = coincidentes.to_frame()
    print(df_coincidentes)
    df_alexistexas = alexistexas.to_frame()

    if isinstance(df_alexistexas, pd.DataFrame):
        print('es df')
    else:
        print('no es df')
    if isinstance(df_coincidentes, pd.DataFrame):
        print('es df')
    else:
        print('no es df')

    print(df_alexistexas)
    pdb.set_trace()
    wb = openpyxl.load_workbook(archivo_sfp_excel)
    ws = wb.active
    # Descombinar todas las celdas en la hoja
    #obtener el formato de relleno de color del DataFrame resultado
    fill_format = {
        True: PatternFill(start_color='0EE1A4', end_color='0EE1A4', fill_type='solid'),  #verde si hay 
        False: PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid'), #rojo si no hay
    }
    for row_idx, row in enumerate(df_alexistexas.values,start=2): #fila
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=row_idx + 3, column=col_idx + 7)
            fill = fill_format.get(df_coincidentes.iat[row_idx -2, col_idx])  # Obtener el formato de relleno de color basado en df
            if fill :
                cell.fill = fill
            cell.value = value
            # Asignar el valor a la celda
            #ws.cell(row=row_idx, column=col_idx+ 2,value=value)
    #ver la forma de copiar los valores con colores en la columna especifica No. de Serie , sobre el archivo Template
    wb.save(archivo_sfp_excel)
    print("Archivo FRAME guardado exitosamente :D")


def abrir_ubicacion(template):
    if template: 
        respuesta = messagebox.askyesno("Confirmar", "¿Desea abrir la carpeta?")
        if respuesta:
            try:
                os.startfile(template)
            except OSError as e:
                messagebox.showerror("Error", f"No se pudo abrir la ubicación de la carpeta:\n{str(e)}")
    else:
        messagebox.showinfo("Información", "No se ha seleccionado una ubicación de carpeta.")


card_report = "D:\\Documentos\\ATP Inventario Tool\\newchanges\\Card Report_2023-10-18_10-21-40_0.xlsx"
sfp_report = "D:\\Documentos\\ATP Inventario Tool\\newchanges\\SFP_Information_Report_2023-10-18_10-23-09.xlsx"
frame_report =  "D:\\Documentos\\ATP Inventario Tool\\newchanges\\Frame Report_2023-10-18_10-18-21_0.xlsx"
#atp_Inventario = "D:\\Documentos\\ATP Inventario Tool\\test1\\LISTA DE MATERIAL DE PER.CENTRO FER.xlsx"
atp_Inventario = "D:\\Documentos\\ATP Inventario Tool\\newchanges\\Book1.xlsx"
#================carlos
#card_report = "D:\\Documentos\\ATP Inventario Tool\\test1\\Card Report_2023-10-18_12-20-19_0 FER.xlsx"
#sfp_report = "D:\\Documentos\\ATP Inventario Tool\\test1\\SFP_Information_Report_2023-10-18_12-21-07 FER.xlsx"
#atp_Inventario = "D:\\Documentos\\ATP Inventario Tool\\test1\\LISTA DE MATERIAL DE PER.CENTRO FER.xlsx"
#atp_Inventario = "D:\\Documentos\\ATP Inventario Tool\\test1\\LISTA DE MATERIAL DE PER.CENTRO FER.xlsx"
#atp_Inventario = "C:\\Users\\j84319062\\Documents\\GitHub\\ATP_Inventario\\input\\Inventario 499P28L+Cun.Villeta+ON.xlsx"

generate_report(card_report,sfp_report,frame_report,atp_Inventario) # ejecuta la funcion para generar reporte
